import os
import time
import pyglet
import winsound
import pandas as pd
import openpyxl
import openpyxl.utils

from threading import Thread, Lock
from gtts import gTTS
from winotify import Notification
from openpyxl.styles import Alignment, Font, PatternFill
from os import PathLike

from app import config

bloqueo_hablar = Lock()


def avisos(msg: str, titulo="", error=False, hablado=True, beep=True):
    notification_win(msg, titulo, error)
    if hablado:
        hilo_hablar = Thread(
            target=hablar,
            kwargs={"msg": msg, "error": error, "beep": beep}
        )
        hilo_hablar.start()


def notification_win(
        msg: str,
        titulo="",
        error=False,
        aplicacion=config.APP_NOMBRE,
        con_voz=False
):
    # nombre_fichero = os.path.join(os.getcwd(), "images\\reus_peq.ico")
    fich_icono = os.path.join(os.getcwd(), "images\\eye.svg")

    cartelito = Notification(
        app_id=aplicacion,
        title=titulo,
        msg=msg,
        duration="long",
        icon=fich_icono
    )
    cartelito.show()

    if con_voz:
        hilo_hablar = Thread(target=hablar,
                             kwargs={"msg": msg, "error": error, "beep": beep})
        hilo_hablar.start()


def hablar(msg: str, beep=False, error=False):
    bloqueo_hablar.acquire()
    if beep:
        _beep(error)

    fichero = os.path.join(os.getcwd(), r"tmp\hablar.mp3")

    tts = gTTS(text=msg, lang="es", tld="es", slow=False)
    if os.path.exists(fichero):
        os.remove(fichero)
    tts.save(fichero)

    sonido = pyglet.media.load(fichero, streaming=False)
    sonido.play()
    time.sleep(sonido.duration)  # prevenir que el delete mate el proceso
    if os.path.exists(fichero):
        os.remove(fichero)

    bloqueo_hablar.release()


def _beep(error=False):
    normal_sound = os.path.join(
        os.getcwd(), "images\\mixkit-melodic-bonus-collect-1938.wav"
    )
    error_sound = os.path.join(
        os.getcwd(), "images\\Error.wav"
    )
    fichero = error_sound if error else normal_sound
    winsound.PlaySound(fichero, winsound.SND_ASYNC)
    time.sleep(1.5)


def exportar_excel(
        fich: str | PathLike,
        data: dict,
        index_excel: bool = False,
        mode: str = "w",
        ancho_columnas: dict = None
) -> None:
    """
        Facilita la exportanción a Excel de Pandas.DataFrames
            fich = Nombre completo del fichero Excel de salida
            data = Es un diccionario cuyas claves son los nombres de ls hojas
                      del Excel a crear y cuyos datos son objetos DataFrame que van
                      a exportarse
            index_excel = bool - Si es True el índice se exportará al Excel.
                      Por defecto es False
            fmode = 'W' para sobreescribir o "a" para añadir al fichero existente
                     pero sobreescribiendo las pestañas pre-existentes
            ancho_columnas = opcional, diccionario con los anchos deseados para las
                             las columnas sobre las que no se prefiera el ancho
                             automático (autofit)
        """
    if not isinstance(data, dict):
        msg = "Parámetro 'data' debe ser de tipo 'dict'..."
        # logger.error(msg)
        raise TypeError(msg)

    # Borrar el fichero si existe y si el modo no es "a"
    if os.path.exists(fich) and mode != "a":
        os.remove(fich)
    else:
        print(f"Escribiendo fichero de salida '{fich}'")

    with pd.ExcelWriter(
            fich, date_format="yyyy-mm-dd", mode=mode, engine="openpyxl",
            if_sheet_exists="replace" if mode == "a" else None
    ) as writer:
        def as_text(value):
            return "" if value is None else str(value)

        for hoja, df in data.items():
            if not isinstance(df, pd.DataFrame):
                msg = "Parámetro 'datos' debe ser de tipo 'pandas.DataFrame'..."
                raise TypeError(msg)

            df.to_excel(writer,
                        sheet_name=hoja,
                        header=True,
                        # engine="openpyxl",
                        index=index_excel,
                        merge_cells=False,
                        # encoding="utf-8",
                        freeze_panes=(1, 0)
                        )

            wb = writer.book
            ws = wb[hoja]
            fill_cabecera = PatternFill(
                fgColor="00858C", fill_type="solid",
                start_color="00858C", end_color="00858C"
            )
            font_cabecera = Font(name="Consolas", size=10, color="FFFFFF",
                                 bold=True)
            font_cuerpo = Font(name="Consolas", size=10, color="000000",
                               bold=False)
            alin_cabecera = Alignment(vertical="bottom")
            alin_cuerpo = Alignment(vertical="top")
            ult_col = len(df.columns) + 1 if index_excel else len(df.columns)
            for row in ws.iter_rows(min_row=1, min_col=1, max_row=len(df) + 1,
                                    max_col=ult_col):
                for cell in row:
                    if cell.row == 1:
                        cell.fill = fill_cabecera
                        cell.font = font_cabecera
                        cell.alignment = alin_cabecera
                    else:
                        cell.font = font_cuerpo
                        cell.alignment = alin_cuerpo

            for column_cells in ws.columns:
                if ancho_columnas and column_cells[0].value in ancho_columnas:
                    new_column_length = ancho_columnas[column_cells[0].value]
                else:
                    new_column_length = max(
                        len(as_text(cell.value)) for cell in column_cells)
                new_column_letter = (openpyxl.utils.get_column_letter(
                    column_cells[0].column))

                if new_column_length > 0:
                    ws.column_dimensions[
                        new_column_letter].width = new_column_length + 1
