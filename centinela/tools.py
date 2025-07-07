import os
import time
import pyglet
import winotify
import openpyxl
import openpyxl.utils
import pandas as pd
import logging

from typing import Union, Literal
from os import PathLike
from pandas import DataFrame
from threading import Thread, Lock
from gtts import gTTS
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

from centinela import config

logger = logging.getLogger(__name__)
bloqueo_hablar = Lock()


def mostrar_notificacion(
        titulo: str = "Atención...",
        msg: str = "Aplicación Centinela, ejecutándose en "
                   "segundo plano desde la bandeja del sistema...",
        sonido=winotify.audio.Reminder,
        msg_hablado: str = ""
):
    toast = winotify.Notification(
        app_id=config.APP_NOMBRE,
        title=titulo,
        msg=msg,
        duration="short",
        icon=config.ICONO_ACTIVO_FICH
    )
    toast.set_audio(sonido, loop=False)
    toast.show()

    if msg_hablado:
        hilo_hablar = Thread(target=hablar, kwargs={"msg": msg_hablado})
        hilo_hablar.start()


def hablar(msg: str):
    bloqueo_hablar.acquire()
    carpeta = os.path.join(Path(__file__).parent, "tmp")
    if not os.path.isdir(carpeta):
        os.mkdir(carpeta)

    fichero = os.path.join(carpeta, "hablar.mp3")
    tts = gTTS(text=msg, lang="es", tld="es", slow=False)
    if os.path.exists(fichero):
        os.remove(fichero)
    tts.save(fichero)

    sonido = pyglet.media.load(fichero, streaming=False)
    sonido.play()
    time.sleep(sonido.duration)  # prevenir que el delete mate el proceso
    if os.path.exists(fichero):
        os.remove(fichero)

    bloqueo_hablar.release()


def exportar_excel_old(
        fich: str | PathLike,
        data: dict,
        index_excel: bool = False,
        mode: Literal["w", "a"] = "w",
        ancho_columnas: dict = None
) -> None:
    if not isinstance(data, dict):
        msg = "Parámetro 'data' debe ser de tipo 'dict'..."
        # logger.error(msg)
        raise TypeError(msg)

    # Borrar el fichero si existe y si el modo no es "a"
    if os.path.exists(fich) and mode != "a":
        os.remove(fich)
    else:
        print(f"Escribiendo fichero de salida '{fich}'")

    with pd.ExcelWriter(
            fich, date_format="yyyy-mm-dd", mode=mode, engine="openpyxl",
            if_sheet_exists="replace" if mode == "a" else None
    ) as writer:
        def as_text(value):
            return "" if value is None else str(value)

        for hoja, df in data.items():
            if not isinstance(df, pd.DataFrame):
                msg = "Parámetro 'datos' debe ser de tipo 'pandas.DataFrame'..."
                raise TypeError(msg)

            df.to_excel(writer,
                        sheet_name=hoja,
                        header=True,
                        # engine="openpyxl",
                        index=index_excel,
                        merge_cells=False,
                        # encoding="utf-8",
                        freeze_panes=(1, 0)
                        )

            wb = writer.book
            ws = wb[hoja]
            fill_cabecera = PatternFill(
                fgColor="00858C", fill_type="solid",
                start_color="00858C", end_color="00858C"
            )
            font_cabecera = Font(name="Consolas", size=10, color="FFFFFF",
                                 bold=True)
            font_cuerpo = Font(name="Consolas", size=10, color="000000",
                               bold=False)
            alin_cabecera = Alignment(vertical="bottom")
            alin_cuerpo = Alignment(vertical="top")
            ult_col = len(df.columns) + 1 if index_excel else len(df.columns)
            for row in ws.iter_rows(min_row=1, min_col=1, max_row=len(df) + 1,
                                    max_col=ult_col):
                for cell in row:
                    if cell.row == 1:
                        cell.fill = fill_cabecera
                        cell.font = font_cabecera
                        cell.alignment = alin_cabecera
                    else:
                        cell.font = font_cuerpo
                        cell.alignment = alin_cuerpo

            for column_cells in ws.columns:
                if ancho_columnas and column_cells[0].value in ancho_columnas:
                    new_column_length = ancho_columnas[column_cells[0].value]
                else:
                    new_column_length = max(
                        len(as_text(cell.value)) for cell in column_cells
                    )
                new_column_letter = (
                    openpyxl.utils.get_column_letter(
                        column_cells[0].column
                    )
                )

                if new_column_length > 0:
                    ws.column_dimensions[
                        new_column_letter].width = new_column_length + 1


# from pathlib import Path
# from typing import Union, Literal
# import os
# import pandas as pd
# from pandas import DataFrame
# from openpyxl.styles import PatternFill, Font, Alignment
# from openpyxl.utils import get_column_letter


def exportar_excel(
        fich: Union[str, Path],
        data: dict[str, DataFrame],
        index_excel: bool = False,
        mode: Literal["w", "a"] = "w",
        ancho_columnas: dict[str, int] = None,
        alineacion_columnas: dict[str, Literal["left", "center", "right"]] = None
) -> None:
    fich = Path(fich)

    if not isinstance(data, dict) or not all(isinstance(df, pd.DataFrame) for df in data.values()):
        raise TypeError("El parámetro 'data' debe ser un diccionario con objetos DataFrame como valores.")

    if mode != "a" and fich.exists():
        fich.unlink()
    else:
        logger.debug(f"Escribiendo fichero de salida '{fich}'")

    with pd.ExcelWriter(fich, date_format="yyyy-mm-dd", mode=mode,
                        engine="openpyxl", if_sheet_exists="replace"
            if mode == "a" else None) as writer:

        for hoja, df in data.items():
            df.to_excel(writer, sheet_name=hoja, header=True,
                        index=index_excel, merge_cells=False,
                        freeze_panes=(1, 0))

            wb = writer.book
            ws = wb[hoja]

            # Estilos
            fill_cabecera = PatternFill(fgColor="00858C", fill_type="solid")
            font_cabecera = Font(name="Consolas", size=10, color="FFFFFF", bold=True)
            font_cuerpo = Font(name="Consolas", size=10, color="000000", bold=False)

            # Crear alineaciones de columnas
            alineaciones = {}
            if alineacion_columnas:
                for col, alin in alineacion_columnas.items():
                    alineaciones[col] = Alignment(horizontal=alin, vertical="top")
            else:
                for col in df.columns:
                    if pd.api.types.is_numeric_dtype(df[col]):
                        alin = "right"
                    else:
                        alin = "left"
                    alineaciones[col] = Alignment(horizontal=alin, vertical="top")

            # Obtener encabezados en orden según fila 1
            headers = [cell.value for cell in ws[1]]

            # Aplicar estilos a cabecera y cuerpo
            for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, max_col=len(headers) + (1 if index_excel else 0)):
                for cell in row:
                    col_idx = cell.column - 1
                    if index_excel:
                        col_idx -= 1
                        if col_idx < 0:
                            continue  # columna del índice

                    col_name = headers[col_idx] if col_idx < len(headers) else None

                    if cell.row == 1:
                        align = alineaciones.get(col_name, Alignment(horizontal="left", vertical="bottom"))
                        cell.fill = fill_cabecera
                        cell.font = font_cabecera
                        cell.alignment = Alignment(horizontal=align.horizontal, vertical="bottom")
                    else:
                        align = alineaciones.get(col_name, Alignment(horizontal="left", vertical="top"))
                        cell.font = font_cuerpo
                        cell.alignment = align

            # Ajuste de anchos de columnas
            def as_text(value):
                return "" if value is None else str(value)

            for column_cells in ws.columns:
                col_name = column_cells[0].value
                if ancho_columnas and col_name in ancho_columnas:
                    new_column_length = ancho_columnas[col_name]
                else:
                    new_column_length = max(len(as_text(cell.value)) for cell in column_cells)
                new_column_letter = get_column_letter(column_cells[0].column)
                if new_column_length > 0:
                    ws.column_dimensions[new_column_letter].width = new_column_length + 1

def reemplazar_en_lista(items: list[str],
                        buscar: list[str],
                        sustituir: list[str]) -> list[str]:
    if len(buscar) != len(sustituir):
        raise ValueError("Las listas 'buscar' y 'sustituir' deben tener la misma longitud")

    resultado = []
    for item in items:
        nuevo = item
        for b, s in zip(buscar, sustituir):
            nuevo = nuevo.replace(b, s)
        resultado.append(nuevo)

    return resultado
